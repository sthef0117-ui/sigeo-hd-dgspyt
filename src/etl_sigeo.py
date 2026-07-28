"""
SIGEO-HD DGSPYT — Pipeline de datos.

Lee los insumos originales (Excel entregados por DGSPYT y C5) y produce los
JSON que consume el tablero, mas las tablas de la base SQLite.

    python src/etl_sigeo.py

Salidas en analisis/:
    corroborados_sigeo.json   56 homicidios dolosos corroborados, julio 2026
    llamadas_911_sigeo.json   llamadas C5 con geocodificacion Windows Maps
    bases_dgspyt.json         inventario de inmuebles con coordenadas y personal
    zonas_ciegas.json         sectores con violencia y sin cobertura de patrullaje
    auditoria_decesos.json    triage de decesos dudosos / suicidios / no localizados
    resumen_ejecutivo.json    KPIs calculados para la reunion de mandos C5

Datos personales: el pipeline NO exporta nombre, apellido ni telefono del
reportante 911. Esas columnas se descartan al leer el Excel.
"""

import json
import math
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from nlp_geocoder import geocodificar, _norm  # noqa: E402
from anonimizar import anonimizar_lista  # noqa: E402

RAIZ = Path(__file__).resolve().parent.parent
INSUMOS = RAIZ / "insumos"
ANALISIS = RAIZ / "analisis"
BD = RAIZ / "database" / "sigeo_db.sqlite"

XLS_HD = INSUMOS / "excel" / "ACCIONES DE HD JULIO DGSPYT.xlsx"
XLS_911 = (INSUMOS / "excel" /
           "LLAMADAS_911_CORTE_DE_LAS_15.00_DEL_26_DE_JULIO_A_LAS_03.00_"
           "DEL_27_DE_JULIO_DE_2026_LIMPIO.xlsx")
XLS_BASES = INSUMOS / "whatsapp" / "INMUEBLES GENERAL DGSPYT.xlsx"

NULOS = {"", "0", "0.0", "NONE", "NULL", "N/A", "NA", "SIN DATO",
         "SIN INFORMACION", "SIN DOCUMENTO", "SIN DOCUMENTACION"}


# --------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------

def limpio(valor):
    if valor is None:
        return ""
    s = re.sub(r"\s+", " ", str(valor)).strip()
    return "" if s.upper() in NULOS else s


def a_float(valor):
    try:
        f = float(str(valor).strip())
    except (TypeError, ValueError):
        return None
    return None if f == 0 else f


def a_hora(valor):
    """Normaliza los formatos de hora heterogeneos de los dos Excel."""
    if valor is None:
        return ""
    if hasattr(valor, "strftime"):
        return valor.strftime("%H:%M:%S")
    s = str(valor).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}:{m.group(3) or '00'}"
    try:  # fraccion de dia de Excel
        frac = float(s)
        if 0 <= frac < 1:
            seg = round(frac * 86400)
            return f"{seg // 3600:02d}:{seg % 3600 // 60:02d}:{seg % 60:02d}"
    except ValueError:
        pass
    return ""


def a_fecha(valor):
    if valor is None:
        return ""
    if hasattr(valor, "strftime"):
        return valor.strftime("%Y-%m-%d")
    s = str(valor).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10]


def minutos(hhmmss):
    m = re.match(r"^(\d{2}):(\d{2})", hhmmss or "")
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def haversine_km(lat1, lng1, lat2, lng2):
    """Distancia ortodromica en km. Misma formula que el motor C++ del proyecto."""
    r = 6371.0088
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = p2 - p1
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


# Los tres insumos escriben el mismo municipio de forma distinta: la tabla de
# HD usa "ECATEPEC", la del 911 "ECATEPEC DE MORELOS" y el inventario de
# inmuebles "Ecatepec de Morelos". Sin unificar, un municipio aparece partido
# en dos y sus homicidios quedan separados de sus propias llamadas.
SUFIJOS_MUNICIPIO = (
    " DE MORELOS", " DE JUAREZ", " DE ZARAGOZA", " DE BAZ", " SOLIDARIDAD",
)
ALIAS_MUNICIPIO = {
    "CIUDAD NEZAHUALCOYOTL": "NEZAHUALCOYOTL",
    "NEZA": "NEZAHUALCOYOTL",
    "CUAUTITLAN MEXICO": "CUAUTITLAN",
}


def municipio_canonico(nombre):
    """Nombre unico de municipio para poder cruzar los tres insumos."""
    n = _norm(nombre)
    if not n or n in NULOS:
        return ""
    n = ALIAS_MUNICIPIO.get(n, n)
    for sufijo in SUFIJOS_MUNICIPIO:
        if n.endswith(sufijo):
            n = n[: -len(sufijo)].strip()
            break
    return n


def en_edomex(lat, lng):
    """Descarta coordenadas fuera del envolvente del Estado de Mexico."""
    return lat is not None and lng is not None and 18.2 <= lat <= 20.4 and -100.7 <= lng <= -98.3


# --------------------------------------------------------------------------
# 1. Homicidios dolosos corroborados (hoja CORROBORADOS)
# --------------------------------------------------------------------------

def cargar_hd():
    wb = openpyxl.load_workbook(XLS_HD, read_only=True, data_only=True)
    ws = wb["CORROBORADOS"]
    registros = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        np_ = limpio(fila[0])
        if not np_ or not np_.replace(".0", "").isdigit():
            continue
        lat, lng = a_float(fila[7]), a_float(fila[8])
        if not en_edomex(lat, lng):
            lat = lng = None
        municipio = municipio_canonico(fila[4])
        colonia = limpio(fila[5])
        calle = limpio(fila[6])
        geo = geocodificar(municipio=municipio, colonia=colonia, calle=calle,
                           notas=limpio(fila[13]), lat=lat, lng=lng)
        registros.append({
            "id": int(float(np_)),
            "fecha": a_fecha(fila[1]),
            "dia_semana": limpio(fila[2]),
            "hora": a_hora(fila[3]),
            "municipio": municipio,
            "colonia": colonia,
            "calle": calle,
            "cuadrante": limpio(fila[12]),
            "lat": lat,
            "lng": lng,
            "total_hd": int(float(limpio(fila[9]) or 1)),
            "sexo": limpio(fila[11]),
            "movil": limpio(fila[14]) or "Se desconoce el móvil de la agresión",
            "observaciones": limpio(fila[10]),
            "desarrollo_hechos": limpio(fila[13]),
            "informacion_adicional": limpio(fila[15]),
            "acciones_ssem": limpio(fila[16]),
            "windows_maps_query": geo["query"],
            "geo_confianza": geo["confianza"],
            "uri_windows_maps": geo["uri_windows_maps"],
        })
    wb.close()
    return registros


# --------------------------------------------------------------------------
# 2. Llamadas de emergencia 911 / C5
# --------------------------------------------------------------------------

# Ponderacion de violencia por palabra clave del catalogo de incidentes C5.
# Se evalua en orden: gana la primera regla que coincide.
REGLAS_VIOLENCIA = [
    (r"HOMICIDIO", 10.0, "letal"),
    (r"LESIONAD.*PROYECTIL|PROYECTIL DE ARMA DE FUEGO", 6.0, "arma_fuego"),
    (r"HUELLAS DE VIOLENCIA", 5.0, "letal"),
    (r"DETONACION.*ARMA DE FUEGO", 4.0, "arma_fuego"),
    (r"PORTACION DE ARMA(S)? DE FUEGO|CARTUCHOS", 3.0, "arma_fuego"),
    (r"LESIONAD.*ARMA BLANCA", 3.0, "arma_blanca"),
    (r"SUICIDIO", 3.0, "deceso_dudoso"),
    (r"PERSONA TIRADA", 3.0, "deceso_dudoso"),
    (r"NO LOCALIZADA|DESAPARECIDA", 3.0, "desaparicion"),
    (r"PORTACION DE ARMA BLANCA", 2.0, "arma_blanca"),
    (r"CON VIOLENCIA", 2.0, "robo_violento"),
    (r"VIOLENCIA CONTRA LAS MUJERES|VIOLENCIA FAMILIAR|VIOLENCIA DE PAREJA",
     2.0, "violencia_genero"),
    (r"RIÑA|RINA|PERSONA AGRESIVA|ALTERACION DEL ORDEN", 1.5, "rina"),
    (r"LESIONAD.*OBJETO CONTUNDENTE", 1.5, "rina"),
    (r"SECUESTRO|PRIVACION DE LA LIBERTAD|EXTORSION", 4.0, "alta_incidencia"),
]


def clasificar_incidente(descripcion):
    d = _norm(descripcion)
    for patron, peso, familia in REGLAS_VIOLENCIA:
        if re.search(patron, d):
            return peso, familia
    return 0.0, "no_violento"


def cargar_911():
    wb = openpyxl.load_workbook(XLS_911, read_only=True, data_only=True)
    ws = wb["REPORTE C.A.LL.E 9-1-1"]
    registros = []
    # Fila 4 = encabezados; los datos inician en la 5.
    # Columnas de reportante (11,12,13) se descartan por proteccion de datos.
    for fila in ws.iter_rows(min_row=5, values_only=True):
        folio = limpio(fila[1])
        if not folio:
            continue
        incidente = limpio(fila[3])
        lat, lng = a_float(fila[28]), a_float(fila[29])
        if not en_edomex(lat, lng):
            lat = lng = None
        municipio = municipio_canonico(fila[9])
        direccion = limpio(fila[10])
        referencia = limpio(fila[27])
        notas = limpio(fila[14])
        peso, familia = clasificar_incidente(incidente)
        geo = geocodificar(municipio=municipio, direccion=direccion,
                           referencia=referencia, notas=notas, lat=lat, lng=lng)
        registros.append({
            "folio": folio,
            "tipo": limpio(fila[2]),
            "incidente": incidente,
            "familia": familia,
            "peso_violencia": peso,
            "fecha": a_fecha(fila[4]),
            "hora": a_hora(fila[6]),
            "municipio": municipio,
            "direccion": direccion,
            "referencia": referencia,
            "notas": notas,
            "corporacion": limpio(fila[23]),
            "modo_recepcion": limpio(fila[25]),
            "codigo_cancelacion": limpio(fila[26]),
            "lat": lat,
            "lng": lng,
            "windows_maps_query": geo["query"],
            "geo_confianza": geo["confianza"],
            "geo_fuentes": geo["fuentes"],
            "geo_componentes": geo["componentes"],
            "uri_windows_maps": geo["uri_windows_maps"],
            "uri_ruta_despacho": geo["uri_ruta_despacho"],
        })
    wb.close()
    return registros


# --------------------------------------------------------------------------
# 3. Inventario de inmuebles / bases DGSPYT
# --------------------------------------------------------------------------

def cargar_bases():
    wb = openpyxl.load_workbook(XLS_BASES, read_only=True, data_only=True)
    ws = wb["DGSPYT"]
    bases = []
    for fila in ws.iter_rows(min_row=7, values_only=True):
        cvo = limpio(fila[0])
        if not cvo or not cvo.replace(".0", "").isdigit():
            continue
        coord = limpio(fila[7])
        m = re.match(r"^\s*(-?\d+\.?\d*)\s*,\s*(-?\d+\.?\d*)\s*$", coord)
        lat = a_float(m.group(1)) if m else None
        lng = a_float(m.group(2)) if m else None
        if not en_edomex(lat, lng):
            lat = lng = None

        def entero(v):
            try:
                return int(float(str(v)))
            except (TypeError, ValueError):
                return 0

        bases.append({
            "cvo": int(float(cvo)),
            "municipio": municipio_canonico(fila[1]),
            "ubicacion": limpio(fila[3]) or limpio(fila[2]),
            "unidad": limpio(fila[5]) or limpio(fila[4]),
            "tipo_construccion": limpio(fila[12]),
            "en_uso": limpio(fila[13]).upper().startswith("SI"),
            "lat": lat,
            "lng": lng,
            "personal_hombres": entero(fila[22]),
            "personal_mujeres": entero(fila[23]),
            "personal_total": entero(fila[24]),
        })
    wb.close()
    return bases


# --------------------------------------------------------------------------
# 4. Detector de zonas ciegas de patrullaje
# --------------------------------------------------------------------------

CELDA_GRADOS = 0.02  # ~2.2 km de lado


def _celda(lat, lng):
    return (math.floor(lat / CELDA_GRADOS), math.floor(lng / CELDA_GRADOS))


def _centro(celda):
    return ((celda[0] + 0.5) * CELDA_GRADOS, (celda[1] + 0.5) * CELDA_GRADOS)


def detectar_zonas_ciegas(hd, llamadas, bases, tope=None):
    """
    "¿Por que no se patrulla ahi?" convertido en una metrica reproducible.

    Para cada celda de ~2.2 km con violencia registrada se calcula:
      indice_violencia  suma ponderada de HD (x10) y llamadas por familia
      dist_base_km      distancia a la base DGSPYT en uso mas cercana
      personal_3km      elementos adscritos a bases dentro de 3 km
      indice_ceguera    violencia x penalizacion por lejania y baja capacidad

    El sector se clasifica ademas en dos diagnosticos que exigen decisiones
    opuestas en la reunion de mandos:

      DESATENDIDO  hay violencia y no hay despliegue cerca -> extender cobertura
      SATURADO     hay despliegue cerca y aun asi concentra violencia ->
                   revisar efectividad del patrullaje existente, no sumar bases

    Un indice alto no afirma que exista omision; senala donde revisar.
    """
    bases_geo = [b for b in bases if b["lat"] and b["en_uso"]]
    celdas = defaultdict(lambda: {
        "eventos_hd": 0, "victimas_hd": 0, "llamadas": 0,
        "indice_violencia": 0.0, "municipios": Counter(),
        "colonias": Counter(), "incidentes": Counter(),
        "familias": Counter(), "hd_detalle": [],
    })

    for d in hd:
        if not (d["lat"] and d["lng"]):
            continue
        c = celdas[_celda(d["lat"], d["lng"])]
        c["eventos_hd"] += 1
        c["victimas_hd"] += d["total_hd"]
        c["indice_violencia"] += 10.0 * d["total_hd"]
        c["municipios"][d["municipio"]] += 1
        if d["colonia"]:
            c["colonias"][d["colonia"]] += 1
        c["hd_detalle"].append({"id": d["id"], "fecha": d["fecha"],
                                "colonia": d["colonia"], "movil": d["movil"]})

    for l in llamadas:
        if not (l["lat"] and l["lng"]) or l["peso_violencia"] <= 0:
            continue
        c = celdas[_celda(l["lat"], l["lng"])]
        c["llamadas"] += 1
        c["indice_violencia"] += l["peso_violencia"]
        c["municipios"][l["municipio"]] += 1
        c["incidentes"][l["incidente"]] += 1
        c["familias"][l["familia"]] += 1

    sectores = []
    for celda, datos in celdas.items():
        if datos["indice_violencia"] < 6:  # ruido de fondo
            continue
        lat, lng = _centro(celda)

        base_cercana, dist_min = None, float("inf")
        personal_3km = personal_1km = 0
        bases_3km = 0
        for b in bases_geo:
            d = haversine_km(lat, lng, b["lat"], b["lng"])
            if d < dist_min:
                dist_min, base_cercana = d, b
            if d <= 3.0:
                personal_3km += b["personal_total"]
                bases_3km += 1
            if d <= 1.0:
                personal_1km += b["personal_total"]

        # Penalizacion por lejania (1.0 a 4.0) y por ausencia de elementos
        # desplegados a 3 km (1.0 a ~0.4). La violencia manda; la cobertura
        # modula. Asi un sector caliente no queda sepultado por celdas rurales
        # con dos llamadas y ninguna base cerca.
        factor_distancia = 1.0 + min(dist_min, 6.0) / 2.0
        factor_capacidad = 1.0 / (1.0 + personal_3km / 150.0)
        indice_ceguera = (datos["indice_violencia"]
                          * factor_distancia * factor_capacidad)

        violencia = datos["indice_violencia"]
        if violencia >= 10 and (dist_min > 3.0 or personal_3km == 0):
            clasificacion = "DESATENDIDO"
            diagnostico = ("Violencia registrada sin despliegue a distancia útil: "
                           "procede extender cobertura o reasignar cuadrante.")
        elif violencia >= 20 and dist_min <= 1.5:
            clasificacion = "SATURADO"
            diagnostico = ("Hay base y personal a menos de 1.5 km y aun así "
                           "concentra violencia: revisar efectividad y horarios "
                           "del patrullaje existente, no sumar inmuebles.")
        else:
            clasificacion = "MIXTO"
            diagnostico = ("Cobertura intermedia: verificar recorridos y tiempos "
                           "de respuesta en el cuadrante.")

        # Cobertura SSEM declarada en los HD de la celda.
        texto_ssem = " ".join(
            _norm(h.get("acciones_ssem", "")) for h in hd
            if h["lat"] and h["lng"] and _celda(h["lat"], h["lng"]) == celda
        )
        sin_camaras = "NO SE LOCALIZAN CAMARAS" in texto_ssem or \
                      "SIN CAMARAS" in texto_ssem

        municipio = datos["municipios"].most_common(1)[0][0] if datos["municipios"] else ""
        sectores.append({
            "sector_id": f"SEC-{celda[0]}_{celda[1]}",
            "lat": round(lat, 5),
            "lng": round(lng, 5),
            "municipio": municipio,
            "municipios_involucrados": [m for m, _ in datos["municipios"].most_common(3)],
            "colonias": [c for c, _ in datos["colonias"].most_common(3)],
            "eventos_hd": datos["eventos_hd"],
            "victimas_hd": datos["victimas_hd"],
            "llamadas_violentas": datos["llamadas"],
            "indice_violencia": round(datos["indice_violencia"], 1),
            "incidentes_top": datos["incidentes"].most_common(4),
            "familias": dict(datos["familias"]),
            "base_cercana": base_cercana["unidad"] if base_cercana else "",
            "base_cercana_municipio": base_cercana["municipio"] if base_cercana else "",
            "dist_base_km": round(dist_min, 2) if base_cercana else None,
            "bases_en_3km": bases_3km,
            "personal_1km": personal_1km,
            "personal_3km": personal_3km,
            "sin_camaras_reportado": sin_camaras,
            "indice_ceguera": round(indice_ceguera, 1),
            "clasificacion": clasificacion,
            "diagnostico": diagnostico,
            "hd_detalle": datos["hd_detalle"],
            "limitrofe": len(datos["municipios"]) > 1,
        })

    # Doble ranking: por brecha de cobertura y por concentracion de violencia.
    sectores.sort(key=lambda s: s["indice_violencia"], reverse=True)
    for i, s in enumerate(sectores, 1):
        s["rank_violencia"] = i
    sectores.sort(key=lambda s: s["indice_ceguera"], reverse=True)
    for i, s in enumerate(sectores, 1):
        s["ranking"] = i
    return sectores[:tope] if tope else sectores


# --------------------------------------------------------------------------
# 5. Auditoria forense de decesos dudosos / suicidios / no localizados
# --------------------------------------------------------------------------

TIPOS_AUDITABLES = [
    "SUICIDIO", "TENTATIVA DE SUICIDIO", "AMENAZA DE SUICIDIO",
    "PERSONA TIRADA EN VIA PUBLICA",
    "PERSONA TIRADA EN LA VIA PUBLICA CON HUELLAS DE VIOLENCIA",
    "PERSONA NO LOCALIZADA O DESAPARECIDA",
    "HOMICIDIO",
]

# Cada indicador es una regla legible: (clave, peso, etiqueta, patron).
# Pesos positivos = motivo para revisar el expediente.
# Pesos negativos = elemento que permite descartar.
INDICADORES = [
    ("huellas_violencia", 30, "El propio catálogo C5 lo clasifica con huellas de violencia",
     r"HUELLAS DE VIOLENCIA"),
    ("arma_fuego", 25, "La cabina registra arma de fuego, disparo o impacto",
     r"ARMA DE FUEGO|DISPARO|BALAZO|PROYECTIL|DETONACION|CASQUILLO"),
    ("ocultamiento", 25, "Indicios de traslado u ocultamiento del cuerpo",
     r"AMARRAD|ATAD|ENVUELT|ENCOBIJAD|EN UNA BOLSA|COBIJA|MALETA|CINTA CANELA|"
     r"ARROJAD|TIRARON EL CUERPO"),
    ("terceros", 20, "Participación de terceros referida por el reportante",
     r"SUJETOS|MASCULINOS ARMADOS|SE LO LLEVARON|SE LA LLEVARON|LEVANTAR|"
     r"LEVANTARON|PRIVAR DE LA LIBERTAD|LO GOLPEARON|LA GOLPEARON|AGRESOR|"
     r"ENCAÑON|ENCANON|AMAGAD|SOMETID|RETENID|CONTRA SU VOLUNTAD|FORCEJE"),
    ("antecedente_violencia", 15,
     "Antecedente de violencia de pareja o familiar en el propio reporte",
     r"LA AGREDE|LO AGREDE|ES MUY AGRESIVO|MUY VIOLENTO|VIOLENCIA FAMILIAR|"
     r"ORDEN DE PROTECCION|EX ?PAREJA|VIOLENCIA DE PAREJA|LA AMENAZ|LO AMENAZ"),
    ("arma_blanca", 15, "Arma blanca o herida punzocortante referida",
     r"ARMA BLANCA|CUCHILL|NAVAJA|MACHETE|PUNZOCORTANTE|APUÑAL"),
    ("sangre_lesiones", 15, "Sangrado o lesiones visibles descritas",
     r"SANGRE|SANGRAND|ENSANGRENTAD|HERIDA|LESION(?!AD[OA] POR CAIDA)|GOLPES"),
    ("sin_signos_vitales", 10, "Persona sin signos vitales al arribo",
     r"SIN SIGNOS VITALES|SIN VIDA|FALLECID|OCCIS|CADAVER"),
    ("nocturno", 5, "Hecho ocurrido en franja nocturna (22:00-06:00)", None),
    ("sin_geo", 5, "Sin coordenada válida: no verificable en campo", None),
    ("sin_notas", 5, "Sin notas de cabina: expediente sin elementos para descartar", None),
    ("proximidad_hd", 20,
     "A menos de 1.5 km de un homicidio doloso corroborado de julio", None),
    ("proximidad_arma", 15,
     "A menos de 1 km y ±6 h de otro reporte por arma de fuego", None),
    # Descarte
    ("localizada", -25, "La persona fue localizada según la propia nota",
     r"SE LOCALIZ|YA FUE LOCALIZAD|APARECIO|YA APARECIO|SE REPORTA LOCALIZAD"),
    ("atencion_medica", -15, "Atención médica documentada con persona consciente",
     r"CONSCIENTE|CRUM|PARAMEDIC|TRASLADAD[OA] AL HOSPITAL|BRINDA ATENCION MEDICA"),
    ("etilico", -10, "Estado etílico referido, patrón habitual de persona tirada",
     r"EBRI|ALCOHOLIZAD|ESTADO DE EBRIEDAD|INTOXICAD"),
    ("sin_lesiones", -10, "La cabina asienta que no presenta lesiones",
     r"SIN LESIONES|NO PRESENTA LESIONES|NIEGA LESIONES"),
]


def auditar_decesos(llamadas, hd):
    """
    Triage documental sobre los decesos dudosos, suicidios y no localizados
    del corte C5, conforme al apunte de la reunion de mandos:

        "Suicidio -> 25... + Identificar, desaparecidos, Homicidios dolosos.
         Investigar si son suicidios."

    IMPORTANTE: el resultado es una PRIORIDAD DE REVISION DOCUMENTAL calculada
    sobre el texto de cabina y la geometria del hecho. No es un dictamen
    pericial ni afirma la mecanica de la muerte. Cada caso lista los
    indicadores exactos que elevaron o bajaron su prioridad.
    """
    tipos_norm = {_norm(t) for t in TIPOS_AUDITABLES}
    cohorte = [l for l in llamadas if _norm(l["incidente"]) in tipos_norm]

    hd_geo = [h for h in hd if h["lat"] and h["lng"]]
    armas = [l for l in llamadas
             if l["lat"] and l["familia"] in ("arma_fuego", "letal")]

    resultados = []
    for caso in cohorte:
        texto = _norm(f"{caso['incidente']} {caso['notas']} {caso['referencia']}")
        puntaje = 0
        activos = []

        for clave, peso, etiqueta, patron in INDICADORES:
            disparado = False
            if patron:
                disparado = bool(re.search(patron, texto))
            elif clave == "nocturno":
                mm = minutos(caso["hora"])
                disparado = mm is not None and (mm >= 22 * 60 or mm < 6 * 60)
            elif clave == "sin_geo":
                disparado = caso["lat"] is None
            elif clave == "sin_notas":
                disparado = not caso["notas"]
            elif clave == "proximidad_hd" and caso["lat"]:
                disparado = any(
                    haversine_km(caso["lat"], caso["lng"], h["lat"], h["lng"]) <= 1.5
                    for h in hd_geo)
            elif clave == "proximidad_arma" and caso["lat"]:
                mm = minutos(caso["hora"])
                for a in armas:
                    if a["folio"] == caso["folio"]:
                        continue
                    if haversine_km(caso["lat"], caso["lng"], a["lat"], a["lng"]) > 1.0:
                        continue
                    ma = minutos(a["hora"])
                    if mm is None or ma is None or abs(mm - ma) <= 360:
                        disparado = True
                        break
            if disparado:
                puntaje += peso
                activos.append({"clave": clave, "peso": peso, "motivo": etiqueta})

        puntaje = max(puntaje, 0)
        if puntaje >= 45:
            nivel = "ALTA"
        elif puntaje >= 20:
            nivel = "MEDIA"
        else:
            nivel = "BAJA"

        if nivel == "ALTA":
            accion = ("Solicitar carpeta a FGJEM y turnar a Criminalística "
                      "para revisión de mecánica de hechos.")
        elif nivel == "MEDIA":
            accion = ("Contrastar nota de cabina contra parte del primer "
                      "respondiente y video del C5.")
        else:
            accion = "Mantener en seguimiento estadístico; sin elementos para reclasificar."

        resultados.append({
            "folio": caso["folio"],
            "incidente": caso["incidente"],
            "municipio": caso["municipio"],
            "fecha": caso["fecha"],
            "hora": caso["hora"],
            "lat": caso["lat"],
            "lng": caso["lng"],
            "windows_maps_query": caso["windows_maps_query"],
            "notas": caso["notas"],
            "puntaje": puntaje,
            "nivel_revision": nivel,
            "indicadores": activos,
            "accion_sugerida": accion,
        })

    orden = {"ALTA": 0, "MEDIA": 1, "BAJA": 2}
    resultados.sort(key=lambda r: (orden[r["nivel_revision"]], -r["puntaje"]))
    return resultados


# --------------------------------------------------------------------------
# 6. Perfil territorial por municipio
# --------------------------------------------------------------------------

def perfil_territorial(hd, llamadas, bases, sectores):
    """
    Ficha por municipio para la mesa con coordinadores territoriales.

    Del apunte de la reunion: "presionar a coordinadores territoriales,
    perfilado de desempeno".

    Importante sobre el alcance: esto mide CONDICIONES DEL TERRITORIO
    (violencia registrada, despliegue disponible, brechas de cobertura), no
    el desempeno individual de una persona. Los insumos no contienen
    asignacion nominal de mando, turnos ni recorridos, asi que atribuir un
    resultado a un coordinador especifico no se sostiene con estos datos.
    Sirve para abrir la conversacion con evidencia, no para calificar gente.
    """
    perfiles = defaultdict(lambda: {
        "hd_eventos": 0, "hd_victimas": 0, "llamadas_violentas": 0,
        "arma_fuego": 0, "bases_en_uso": 0, "personal_total": 0,
        "sectores_desatendidos": 0, "sectores_saturados": 0,
        "distancias": [], "colonias_hd": Counter(), "hd_nocturnos": 0,
    })

    for h in hd:
        p = perfiles[h["municipio"]]
        p["hd_eventos"] += 1
        p["hd_victimas"] += h["total_hd"]
        if h["colonia"]:
            p["colonias_hd"][h["colonia"]] += 1
        mm = minutos(h["hora"])
        if mm is not None and (mm >= 22 * 60 or mm < 6 * 60):
            p["hd_nocturnos"] += 1

    for l in llamadas:
        if l["peso_violencia"] <= 0 or not l["municipio"]:
            continue
        p = perfiles[l["municipio"]]
        p["llamadas_violentas"] += 1
        if l["familia"] == "arma_fuego":
            p["arma_fuego"] += 1

    # El inventario de inmuebles escribe el municipio en otra caja tipografica
    # y a veces con sufijo ("Ecatepec de Morelos" vs "ECATEPEC").
    def clave(nombre):
        return _norm(nombre).replace(" DE MORELOS", "").replace(" DE JUAREZ", "") \
                            .replace(" DE ZARAGOZA", "").replace(" DE BAZ", "").strip()

    indice_bases = defaultdict(list)
    for b in bases:
        if b["en_uso"]:
            indice_bases[clave(b["municipio"])].append(b)

    bases_geo = [b for b in bases if b["lat"] and b["en_uso"]]
    for h in hd:
        if not h["lat"]:
            continue
        d = min((haversine_km(h["lat"], h["lng"], b["lat"], b["lng"])
                 for b in bases_geo), default=None)
        if d is not None:
            perfiles[h["municipio"]]["distancias"].append(d)

    for s in sectores:
        if s["clasificacion"] == "DESATENDIDO":
            perfiles[s["municipio"]]["sectores_desatendidos"] += 1
        elif s["clasificacion"] == "SATURADO":
            perfiles[s["municipio"]]["sectores_saturados"] += 1

    salida = []
    for municipio, p in perfiles.items():
        if not municipio:
            continue
        propias = indice_bases.get(municipio, [])
        p["bases_en_uso"] = len(propias)
        p["personal_total"] = sum(b["personal_total"] for b in propias)

        carga = p["hd_eventos"] * 10 + p["llamadas_violentas"]
        dist_media = (round(sum(p["distancias"]) / len(p["distancias"]), 2)
                      if p["distancias"] else None)
        hd_lejanos = sum(1 for d in p["distancias"] if d > 3)

        # Indice de presion: carga de violencia, penalizada por lejania del
        # despliegue y por sectores sin atender. Los tres componentes se
        # muestran en la ficha para que el numero sea discutible.
        factor_dist = 1.0 + (min(dist_media, 6.0) / 4.0 if dist_media else 0.5)
        factor_gap = 1.0 + p["sectores_desatendidos"] * 0.25
        indice_presion = round(carga * factor_dist * factor_gap, 1)

        salida.append({
            "municipio": municipio,
            "hd_eventos": p["hd_eventos"],
            "hd_victimas": p["hd_victimas"],
            "hd_nocturnos": p["hd_nocturnos"],
            "llamadas_violentas": p["llamadas_violentas"],
            "arma_fuego": p["arma_fuego"],
            "carga_violencia": carga,
            "bases_en_uso": p["bases_en_uso"],
            "personal_total": p["personal_total"],
            "dist_media_hd_base_km": dist_media,
            "hd_a_mas_de_3km": hd_lejanos,
            "sectores_desatendidos": p["sectores_desatendidos"],
            "sectores_saturados": p["sectores_saturados"],
            "colonias_reincidentes": [c for c, n in p["colonias_hd"].most_common(3) if n > 1],
            "indice_presion": indice_presion,
        })

    salida.sort(key=lambda x: x["indice_presion"], reverse=True)
    for i, x in enumerate(salida, 1):
        x["ranking"] = i
    return salida


# --------------------------------------------------------------------------
# 7. Resumen ejecutivo calculado
# --------------------------------------------------------------------------

def resumen_ejecutivo(hd, llamadas, bases, sectores, auditoria):
    hd_geo = [h for h in hd if h["lat"]]
    violentas = [l for l in llamadas if l["peso_violencia"] > 0]
    arma_fuego = [l for l in llamadas if l["familia"] == "arma_fuego"]
    bases_geo = [b for b in bases if b["lat"] and b["en_uso"]]

    # Distancia de cada HD a la base DGSPYT en uso mas cercana.
    distancias = []
    for h in hd_geo:
        d = min((haversine_km(h["lat"], h["lng"], b["lat"], b["lng"])
                 for b in bases_geo), default=None)
        if d is not None:
            distancias.append(d)
    distancias.sort()

    def mediana(xs):
        if not xs:
            return None
        n = len(xs)
        return xs[n // 2] if n % 2 else (xs[n // 2 - 1] + xs[n // 2]) / 2

    municipios_hd = Counter(h["municipio"] for h in hd)
    moviles = Counter(h["movil"] for h in hd)
    # La honestidad del KPI esta en el denominador: la mayoria de las llamadas
    # sin coordenada son mudas, colgadas o de broma y no contienen ningun texto
    # de ubicacion. El universo recuperable son las que si traen descripcion,
    # que es exactamente el problema planteado en la reunion de mandos.
    confianza = Counter(l["geo_confianza"] for l in llamadas)
    sin_coord = [l for l in llamadas if l["lat"] is None]
    recuperables = [l for l in sin_coord
                    if l["direccion"] or l["referencia"] or l["notas"]]
    rescatadas = [l for l in recuperables if l["geo_confianza"] in ("ALTA", "MEDIA")]
    viol_sin_coord = [l for l in sin_coord if l["peso_violencia"] > 0]
    viol_rescatadas = [l for l in viol_sin_coord
                       if l["geo_confianza"] in ("ALTA", "MEDIA")]

    return {
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "corte": "Homicidios dolosos julio 2026 · Llamadas C5 26-27 julio 2026",
        "hd": {
            "eventos": len(hd),
            "victimas": sum(h["total_hd"] for h in hd),
            "georreferenciados": len(hd_geo),
            "sin_coordenada": len(hd) - len(hd_geo),
            "municipios": len(municipios_hd),
            "top_municipios": municipios_hd.most_common(6),
            "top_moviles": moviles.most_common(5),
            "nocturnos": sum(1 for h in hd
                             if (minutos(h["hora"]) or 0) >= 22 * 60
                             or (minutos(h["hora"]) or 0) < 6 * 60),
        },
        "c5": {
            "llamadas_totales": len(llamadas),
            "llamadas_violentas": len(violentas),
            "arma_fuego": len(arma_fuego),
            "con_coordenada": sum(1 for l in llamadas if l["lat"]),
            "top_incidentes_violentos": Counter(
                l["incidente"] for l in violentas).most_common(8),
        },
        "geocodificacion": {
            "distribucion_confianza": dict(confianza),
            "sin_coordenada_original": len(sin_coord),
            "sin_coordenada_sin_texto": len(sin_coord) - len(recuperables),
            "universo_recuperable": len(recuperables),
            "recuperadas_por_nlp": len(rescatadas),
            "porcentaje_recuperado": round(
                100 * len(rescatadas) / max(1, len(recuperables)), 1),
            "violentas_sin_coordenada": len(viol_sin_coord),
            "violentas_recuperadas": len(viol_rescatadas),
            "porcentaje_violentas_recuperado": round(
                100 * len(viol_rescatadas) / max(1, len(viol_sin_coord)), 1),
        },
        "cobertura": {
            "bases_inventariadas": len(bases),
            "bases_en_uso_georreferenciadas": len(bases_geo),
            "personal_total": sum(b["personal_total"] for b in bases if b["en_uso"]),
            "dist_hd_base_promedio_km": round(sum(distancias) / len(distancias), 2)
            if distancias else None,
            "dist_hd_base_mediana_km": round(mediana(distancias), 2) if distancias else None,
            "dist_hd_base_maxima_km": round(max(distancias), 2) if distancias else None,
            "hd_a_mas_de_3km_de_base": sum(1 for d in distancias if d > 3),
        },
        "zonas_ciegas": {
            "sectores_detectados": len(sectores),
            "top": [{"sector_id": s["sector_id"], "municipio": s["municipio"],
                     "indice_ceguera": s["indice_ceguera"],
                     "dist_base_km": s["dist_base_km"],
                     "eventos_hd": s["eventos_hd"],
                     "llamadas_violentas": s["llamadas_violentas"]}
                    for s in sectores[:5]],
        },
        "auditoria": {
            "casos_revisados": len(auditoria),
            "prioridad_alta": sum(1 for a in auditoria if a["nivel_revision"] == "ALTA"),
            "prioridad_media": sum(1 for a in auditoria if a["nivel_revision"] == "MEDIA"),
            "prioridad_baja": sum(1 for a in auditoria if a["nivel_revision"] == "BAJA"),
            "por_tipo": Counter(a["incidente"] for a in auditoria).most_common(),
        },
    }


# --------------------------------------------------------------------------
# 7. Persistencia
# --------------------------------------------------------------------------

def escribir_json(nombre, datos):
    ANALISIS.mkdir(parents=True, exist_ok=True)
    ruta = ANALISIS / nombre
    ruta.write_text(json.dumps(datos, ensure_ascii=False, indent=1), encoding="utf-8")
    return ruta


def escribir_sqlite(hd, llamadas, bases, sectores, auditoria):
    BD.parent.mkdir(parents=True, exist_ok=True)
    cx = sqlite3.connect(BD)
    cur = cx.cursor()

    cur.executescript("""
        DROP TABLE IF EXISTS tb_bases_dgspyt;
        CREATE TABLE tb_bases_dgspyt (
            cvo INTEGER PRIMARY KEY, municipio TEXT, unidad TEXT, ubicacion TEXT,
            tipo_construccion TEXT, en_uso INTEGER, latitud REAL, longitud REAL,
            personal_hombres INTEGER, personal_mujeres INTEGER, personal_total INTEGER
        );
        DROP TABLE IF EXISTS tb_zonas_ciegas;
        CREATE TABLE tb_zonas_ciegas (
            sector_id TEXT PRIMARY KEY, ranking INTEGER, municipio TEXT,
            latitud REAL, longitud REAL, eventos_hd INTEGER, victimas_hd INTEGER,
            llamadas_violentas INTEGER, indice_violencia REAL, dist_base_km REAL,
            base_cercana TEXT, bases_en_3km INTEGER, personal_3km INTEGER,
            indice_ceguera REAL, limitrofe INTEGER
        );
        DROP TABLE IF EXISTS tb_auditoria_decesos_suicidios;
        CREATE TABLE tb_auditoria_decesos_suicidios (
            folio TEXT PRIMARY KEY, incidente TEXT, municipio TEXT, fecha TEXT,
            hora TEXT, latitud REAL, longitud REAL, puntaje INTEGER,
            nivel_revision TEXT, indicadores TEXT, accion_sugerida TEXT
        );
        DROP TABLE IF EXISTS tb_geocodificacion_maps;
        CREATE TABLE tb_geocodificacion_maps (
            folio TEXT PRIMARY KEY, municipio TEXT, direccion_c5 TEXT,
            referencia TEXT, query_windows_maps TEXT, confianza TEXT,
            fuentes TEXT, uri_windows_maps TEXT, latitud REAL, longitud REAL
        );
    """)

    cur.executemany(
        "INSERT INTO tb_bases_dgspyt VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        [(b["cvo"], b["municipio"], b["unidad"], b["ubicacion"],
          b["tipo_construccion"], int(b["en_uso"]), b["lat"], b["lng"],
          b["personal_hombres"], b["personal_mujeres"], b["personal_total"])
         for b in bases])

    cur.executemany(
        "INSERT INTO tb_zonas_ciegas VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(s["sector_id"], s["ranking"], s["municipio"], s["lat"], s["lng"],
          s["eventos_hd"], s["victimas_hd"], s["llamadas_violentas"],
          s["indice_violencia"], s["dist_base_km"], s["base_cercana"],
          s["bases_en_3km"], s["personal_3km"], s["indice_ceguera"],
          int(s["limitrofe"])) for s in sectores])

    cur.executemany(
        "INSERT INTO tb_auditoria_decesos_suicidios VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        [(a["folio"], a["incidente"], a["municipio"], a["fecha"], a["hora"],
          a["lat"], a["lng"], a["puntaje"], a["nivel_revision"],
          json.dumps([i["clave"] for i in a["indicadores"]], ensure_ascii=False),
          a["accion_sugerida"]) for a in auditoria])

    cur.executemany(
        "INSERT INTO tb_geocodificacion_maps VALUES (?,?,?,?,?,?,?,?,?,?)",
        [(l["folio"], l["municipio"], l["direccion"], l["referencia"],
          l["windows_maps_query"], l["geo_confianza"],
          ",".join(l["geo_fuentes"]), l["uri_windows_maps"], l["lat"], l["lng"])
         for l in llamadas])

    # Refresco de las tablas base ya existentes.
    cur.execute("DELETE FROM tb_homicidios_corroborados")
    cur.executemany(
        "INSERT INTO tb_homicidios_corroborados (id, np_consecutivo, fecha_evento,"
        " hora_evento, dia_semana, municipio, colonia, calle, cuadrante, latitud,"
        " longitud, total_victimas_hd, sexo_victima, posible_movil,"
        " desarrollo_hechos, observaciones, acciones_ssem, windows_maps_query,"
        " fecha_registro) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))",
        [(h["id"], h["id"], h["fecha"], h["hora"], h["dia_semana"], h["municipio"],
          h["colonia"], h["calle"], h["cuadrante"], h["lat"], h["lng"],
          h["total_hd"], h["sexo"], h["movil"], h["desarrollo_hechos"],
          h["observaciones"], h["acciones_ssem"], h["windows_maps_query"])
         for h in hd])

    cur.execute("DELETE FROM tb_llamadas_911_c5")
    cur.executemany(
        "INSERT INTO tb_llamadas_911_c5 (folio_c5, tipo_incidente_id,"
        " incidente_descripcion, fecha_llamada, hora_llamada, municipio,"
        " direccion_reportada, referencia_ubicacion, notas_cabina_c5,"
        " modo_recepcion, latitud, longitud, prioridad_nivel, fecha_registro)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))",
        [(l["folio"], l["tipo"], l["incidente"], l["fecha"], l["hora"],
          l["municipio"], l["direccion"], l["referencia"], l["notas"],
          l["modo_recepcion"], l["lat"], l["lng"], l["peso_violencia"])
         for l in llamadas])

    cx.commit()
    cx.close()


# --------------------------------------------------------------------------

def main():
    # Los JSON se publican en GitHub Pages: se anonimizan salvo indicacion
    # expresa. La base SQLite local siempre conserva el texto integro.
    anonimo = "--sin-anonimizar" not in sys.argv

    print("SIGEO-HD DGSPYT · pipeline de datos")
    hd = cargar_hd()
    print(f"  homicidios corroborados .... {len(hd)}")
    llamadas = cargar_911()
    print(f"  llamadas C5 ................ {len(llamadas)}")
    bases = cargar_bases()
    geo_bases = sum(1 for b in bases if b['lat'])
    print(f"  inmuebles DGSPYT ........... {len(bases)} ({geo_bases} georreferenciados)")

    sectores = detectar_zonas_ciegas(hd, llamadas, bases)
    print(f"  sectores de zona ciega ..... {len(sectores)}")
    auditoria = auditar_decesos(llamadas, hd)
    altas = sum(1 for a in auditoria if a['nivel_revision'] == 'ALTA')
    print(f"  casos auditados ............ {len(auditoria)} ({altas} prioridad alta)")

    territorio = perfil_territorial(hd, llamadas, bases, sectores)
    print(f"  municipios perfilados ...... {len(territorio)}")

    resumen = resumen_ejecutivo(hd, llamadas, bases, sectores, auditoria)
    resumen["territorio_top"] = [
        {k: t[k] for k in ("municipio", "indice_presion", "hd_eventos",
                           "llamadas_violentas", "sectores_desatendidos",
                           "personal_total")}
        for t in territorio[:6]]

    # El tablero se distribuye como archivo unico (se abre desde USB en la
    # reunion de mandos), asi que la carga util va recortada: solo llamadas con
    # relevancia operativa y solo los campos que el tablero dibuja.
    CAMPOS_TABLERO = (
        "folio", "incidente", "familia", "peso_violencia", "fecha", "hora",
        "municipio", "direccion", "referencia", "lat", "lng",
        "windows_maps_query", "geo_confianza", "geo_fuentes", "uri_windows_maps",
    )
    llamadas_tablero = []
    for l in llamadas:
        relevante = (l["peso_violencia"] > 0 or l["lat"] is not None
                     or l["geo_confianza"] in ("ALTA", "MEDIA"))
        if not relevante:
            continue
        fila = {c: l[c] for c in CAMPOS_TABLERO}
        # La nota completa vive en SQLite; el tablero solo muestra el extracto.
        nota = l["notas"]
        fila["notas"] = nota[:500] + "…" if len(nota) > 500 else nota
        llamadas_tablero.append(fila)

    def publicable(registros):
        return anonimizar_lista(registros) if anonimo else registros

    resumen["anonimizado"] = anonimo
    escribir_json("corroborados_sigeo.json", publicable(hd))
    escribir_json("llamadas_911_sigeo.json", publicable(llamadas_tablero))
    escribir_json("bases_dgspyt.json", bases)
    escribir_json("zonas_ciegas.json", sectores)
    escribir_json("perfil_territorial.json", territorio)
    escribir_json("auditoria_decesos.json", publicable(auditoria))
    escribir_json("resumen_ejecutivo.json", resumen)
    escribir_sqlite(hd, llamadas, bases, sectores, auditoria)

    print("  datos personales ........... "
          + ("suprimidos en los JSON publicables"
             if anonimo else "SIN SUPRIMIR (--sin-anonimizar)"))

    print(f"  JSON escritos en ........... {ANALISIS}")
    print(f"  SQLite actualizado ......... {BD}")
    g = resumen["geocodificacion"]
    print(f"  NLP recupera {g['recuperadas_por_nlp']} de {g['universo_recuperable']} "
          f"llamadas sin coordenada pero con descripción ({g['porcentaje_recuperado']}%)")
    print(f"  de las violentas sin coordenada: {g['violentas_recuperadas']}"
          f"/{g['violentas_sin_coordenada']} ({g['porcentaje_violentas_recuperado']}%)")


if __name__ == "__main__":
    main()
